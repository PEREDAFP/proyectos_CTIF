#Es necesario tener la hoja cálculo Inscritos.xlsx para realizar la copia

import sys
import shutil
from openpyxl import load_workbook
import collections

def obtenemos_datos():
	'''
	argv[1]---> número de inscritos
	argv[2]---> datos obtenidos de la plataforma de inscritos
	argv[3]---> código del curso en el CTIF
	'''
	if len(sys.argv) < 4:
		print("Debes pasar como parámetros: <número de inscritos> <fichero_datos_inscritos>  <código del curso>")
		exit()
	else:
		return (sys.argv[1],sys.argv[2], sys.argv[3])
		
		
def grabamos_num_inscritos(fichero, num_inscritos):
	fich = load_workbook(fichero)
	hoja = fich['CONFIGURACIÓN']
	#E1 en la hoja CONFIGURACIÓN es donde se guarda el número de inscritos para utilizarlo en los distintos listados
	hoja['e1']=int(num_inscritos)
	fich.save(fichero)


def copiamos_datos_ordenados(fichero, ficherofinal):
	
	ordenA = ['Funcionario interino', 'Funcionario de carrera','Funcionario en prácticas']
	ordenB = ['Contratado laboral (C. Público)']
	ordenC = ['Integrante en listas de interinidad']
	datos = dict()
	fich = load_workbook(fichero)
	hoja = fich['Hoja1']
	celda = "A"
	orden = 2
	while hoja[celda+str(orden)].value != None:
		clave = ""
		#Lo primero...la situación administrativa. Ojo con Religión que está en la columan Q 
		if hoja["q"+str(orden)].value == "Religión":
			clave = "A"
		elif hoja["M" +str(orden)].value in ordenA:
			clave = "A"
		elif hoja["M"+str(orden)].value in ordenB:
			clave = "B"
		elif hoja["M"+str(orden)].value in ordenC:
			clave = "C" 
		#Ahora añadimos el id
		clave += str(hoja['A'+str(orden)].value)
		#Y ahora todo
		fila=[]
		for letra in 'abcdefghijklmnopqrstuvw':
			fila.append(hoja[letra + str(orden)].value)
		datos[clave]=fila
		orden += 1
	#fich = fich.close()	
	#Ahora grabamos en el fichero final hoja INSCRITOS
	fich2 = load_workbook(ficherofinal)
	hoja2 = fich2['INSCRITOS']
	#creamos la lista de letras

	letras = [ x for x in 'abcdefghijklmnopqrstuvwxyz']
	datos_ordenados =collections.OrderedDict(datos)
	orden = 2
	for i in sorted(datos_ordenados.keys()):
		letra = 0
		for dato in datos_ordenados[i]:
			celda = f'{letras[letra]}{orden}'
			hoja2[celda] = dato	
			letra += 1
		orden += 1
		
	fich2.save(ficherofinal)	
			

def main(args):
	
	#Obtenemos los datos de la línea de caracteres
	limite_inscritos, fichero_inscritos, curso = obtenemos_datos()
	
	#Creamos el fichero destino Inscritos_<curso>.xlsx
	fich_final = "Inscritos_"+curso+'.xlsx'
	try:
		shutil.copy("Inscritos.xlsx", fich_final)
	except:
		print("Falta el fichero original de Inscritos.xlsx")
		exit(1)
	
	#Se modifica la celda que contiene el número de inscritos al curso
	grabamos_num_inscritos(fich_final, limite_inscritos)
	
	copiamos_datos_ordenados(fichero_inscritos,fich_final)
	
	
	return 0
	
	 

if __name__ == '__main__':
    import sys
    sys.exit(main(sys.argv))
