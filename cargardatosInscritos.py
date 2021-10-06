#Es necesario tener la hoja cálculo Inscritos.xlsx para realizar la copia

import sys
import shutil
from openpyxl import load_workbook
import collections


LINEACOMIENZO_ORIGEN = 8
LINEACOMIENZO_DESTINO = 2
CELDA_NOMBRE_CURSO = 'A3'

def obtenemos_datos():
	'''
	argv[1]---> número de inscritos
	argv[2]---> datos obtenidos de la plataforma de inscritos
	argv[3]---> código del curso en el CTIF
	argv[4]---> fecha inicio del curso
	'''
	if len(sys.argv) < 5:
		print("Debes pasar como parámetros: <número de inscritos> <fichero_datos_inscritos>  <código del curso> <fecha inicio>")
		exit()
	else:
		return (sys.argv[1],sys.argv[2], sys.argv[3], sys.argv[4])
		
		
def grabamos_configuracion(fichero, num_inscritos, fecha_comienzo):
	#Obtenemos el número de inscritos, el comienzo del curso y el número de participantes
	fich = load_workbook(fichero)
	hoja = fich['CONFIGURACIÓN']
	#E1 en la hoja CONFIGURACIÓN es donde se guarda el número de inscritos para utilizarlo en los distintos listados
	hoja['e1']=int(num_inscritos)
	hoja['e2']=str(fecha_comienzo)
	fich.save(fichero)


def copiamos_datos_ordenados(fichero, ficherofinal):
	
	ordenA = ['Funcionario interino', 'Funcionario de carrera','Funcionario en prácticas']
	ordenB = ['Contratado laboral (C. Público)', 'Contratado temporal (C. Concertado)','Contratado fijo (C. Concertado)']
	ordenC = ['Integrante en listas de interinidad']
	datos = dict()
	fich = load_workbook(fichero)
	hoja = fich['Hoja1']
	celda = "A"
	#Obtenemos el nombre del curso
	nombre_curso = hoja[CELDA_NOMBRE_CURSO].value
	nombre_curso = nombre_curso.split('-')[1].strip() 
	
	
	#Fila desde la que comienza la hoja de cálculo
	orden = LINEACOMIENZO_ORIGEN
	while hoja[celda+str(orden)].value != None:
		clave = ""
		#Lo primero...la situación administrativa. Ojo con Religión que está en la columan Q 
		if hoja["q"+str(orden)].value == "Religión":
			clave = "A"
		elif hoja["M" +str(orden)].value in ordenA:
			clave = "A"
		elif hoja["M"+str(orden)].value in ordenB:
			clave = "B"
		elif hoja["M"+str(orden)].value in ordenC:
			clave = "C" 
		#Ahora añadimos el id
		clave += str(hoja['A'+str(orden)].value)
		#Y ahora todo
		fila=[]
		for letra in 'abcdefghijklmnopqrstuvw':
			fila.append(hoja[letra + str(orden)].value)
		datos[clave]=fila
		orden += 1
	#fich = fich.close()	
	#Ahora grabamos en el fichero final hoja INSCRITOS
	fich2 = load_workbook(ficherofinal)
	hoja2 = fich2['INSCRITOS']
	#creamos la lista de letras

	letras = [ x for x in 'abcdefghijklmnopqrstuvwxyz']
	datos_ordenados =collections.OrderedDict(datos)
	orden = LINEACOMIENZO_DESTINO
	for i in sorted(datos_ordenados.keys()):
		letra = 0
		for dato in datos_ordenados[i]:
			celda = f'{letras[letra]}{orden}'
			hoja2[celda] = dato	
			letra += 1
		orden += 1
	hoja3 = fich2['CONFIGURACIÓN']
	hoja3['e3'] = nombre_curso
		
	fich2.save(ficherofinal)	
			

def main(args):
	
	#Obtenemos los datos de la línea de caracteres
	limite_inscritos, fichero_inscritos, curso, fecha_comienzo = obtenemos_datos()
	
	#Creamos el fichero destino Inscritos_<curso>.xlsx
	fich_final = "Inscritos_"+curso+'.xlsx'
	try:
		shutil.copy("Inscritos.xlsx", fich_final)
	except:
		print("Falta el fichero original de Inscritos.xlsx")
		exit(1)
	
	#Se modifica la celda que contiene el número de inscritos al curso
	grabamos_configuracion(fich_final, limite_inscritos, fecha_comienzo)
	
	copiamos_datos_ordenados(fichero_inscritos,fich_final)
	
	
	return 0
	
	 

if __name__ == '__main__':
    import sys
    sys.exit(main(sys.argv))
