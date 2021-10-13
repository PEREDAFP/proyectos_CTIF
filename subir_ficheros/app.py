# Importamos todo lo necesario

import sys
import shutil
from openpyxl import load_workbook
import collections
import os
from os.path import abspath, dirname, join
from flask import Flask, render_template, request, redirect, url_for
from werkzeug import secure_filename
from forms import File_Form
# instancia del objeto Flask
app = Flask(__name__)

app.config['SECRET_KEY'] = '70c8ae51a4b5af97be6534caef90e4bb9bdcb3380af008f90b23a5d1616bf319bc298105da20fe'
BASE_DIR = dirname(dirname(abspath(__file__)))
# Media dir
MEDIA_DIR = join(BASE_DIR, 'subir_fichero/static/media/')
POSTS_FILES_DIR = join(MEDIA_DIR, 'posts')

# Carpeta de subida
app.config['UPLOAD_FOLDER'] = MEDIA_DIR




#Para los ficheros descargados

LINEACOMIENZO_ORIGEN = 8
LINEACOMIENZO_DESTINO = 2
CELDA_NOMBRE_CURSO = 'A3'

def grabamos_configuracion(fichero, num_inscritos, fecha_comienzo):
	#Obtenemos el número de inscritos, el comienzo del curso y el número de participantes
	fich = load_workbook(fichero)
	hoja = fich['CONFIGURACIÓN']
	#E1 en la hoja CONFIGURACIÓN es donde se guarda el número de inscritos para utilizarlo en los distintos listados
	hoja['e1']=int(num_inscritos)
	hoja['e2']=str(fecha_comienzo)
	fich.save(fichero)


def copiamos_datos_ordenados(fichero, ficherofinal):

	ordenA = ['Funcionario interino', 'Funcionario de carrera','Funcionario en prácticas']
	ordenB = ['Contratado laboral (C. Público)', 'Contratado temporal (C. Concertado)','Contratado fijo (C. Concertado)']
	ordenC = ['Integrante en listas de interinidad']
	datos = dict()
	fich = load_workbook(fichero)
	hoja = fich['Hoja1']
	celda = "A"
	#Obtenemos el nombre del curso
	nombre_curso = hoja[CELDA_NOMBRE_CURSO].value
	nombre_curso = nombre_curso.split('-')[1].strip()


	#Fila desde la que comienza la hoja de cálculo
	orden = LINEACOMIENZO_ORIGEN
	while hoja[celda+str(orden)].value != None:
		clave = ""
		#Lo primero...la situación administrativa. Ojo con Religión que está en la columan Q
		if hoja["q"+str(orden)].value == "Religión":
			clave = "A"
		elif hoja["M" +str(orden)].value in ordenA:
			clave = "A"
		elif hoja["M"+str(orden)].value in ordenB:
			clave = "B"
		elif hoja["M"+str(orden)].value in ordenC:
			clave = "C"
		#Ahora añadimos el id
		clave += str(hoja['A'+str(orden)].value)
		#Y ahora todo
		fila=[]
		for letra in 'abcdefghijklmnopqrstuvw':
			fila.append(hoja[letra + str(orden)].value)
		datos[clave]=fila
		orden += 1
	#fich = fich.close()
	#Ahora grabamos en el fichero final hoja INSCRITOS
	fich2 = load_workbook(ficherofinal)
	hoja2 = fich2['INSCRITOS']
	#creamos la lista de letras

	letras = [ x for x in 'abcdefghijklmnopqrstuvwxyz']
	datos_ordenados =collections.OrderedDict(datos)
	orden = LINEACOMIENZO_DESTINO
	for i in sorted(datos_ordenados.keys()):
		letra = 0
		for dato in datos_ordenados[i]:
			celda = f'{letras[letra]}{orden}'
			hoja2[celda] = dato
			letra += 1
		orden += 1
	hoja3 = fich2['CONFIGURACIÓN']
	hoja3['e3'] = nombre_curso

	fich2.save(ficherofinal)















@app.route('/descarga/<file>', methods=['GET','POST'])
def downloader(fichero):
    return send_file(app.config['UPLOAD_FOLDER'] + fichero, as_attachment=True)

@app.route("/", methods=['POST','GET'])
def uploader():
    form = File_Form()
    if form.validate_on_submit():
        num_curso = form.num_curso.data
        num_matriculados = form.num_matriculados.data
        fecha_inicio = form.fecha_inicio.data
        fichero = form.fichero.data

        if fichero:
            file_name = secure_filename(fichero.filename)
            images_dir = POSTS_FILES_DIR
            os.makedirs(images_dir, exist_ok=True)
            file_path = os.path.join(images_dir, file_name)
            fichero.save(file_path)

            #Ahora empieza a trabajar la grabación de datos...
            fich_final = f'Inscritos_{num_curso}.xlsx'
            try:
                #Obtenemos el fichero Inscritos vacío
                shutil.copy(MEDIA_DIR+"Inscritos.xlsx", f'{MEDIA_DIR}{fich_final}')
            except:
                print("Falta el fichero original de Inscritos.xlsx")
                exit(1)

        next = request.args.get('next', None)
        if next:
            return redirect(next)
        #return redirect(url_for('downloader', file = fich_final))
        #return redirect(fich_final)
        return render_template("download_file.html",fichero = f'media/{fich_final}')
    return render_template("file_form.html", form=form)

if __name__ == '__main__':
 # Iniciamos la aplicación
 app.run(debug=True)

